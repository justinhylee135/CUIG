from __future__ import annotations

import argparse
import csv
import io
import json
import os
import struct
import zipfile
import zlib
from collections import Counter, defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, List, Sequence

import torch
from PIL import Image
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from torch import nn
from torch.utils.data import DataLoader, Dataset
from torchvision import models, transforms
from tqdm import tqdm


DEFAULT_MODEL = "resnet50_copyright_101_71.pt"
DEFAULT_LABELS = "labels.csv"
SUPPORTED_EXTENSIONS = (".jpg", ".jpeg", ".png", ".webp")


def normalize_name(value: str) -> str:
    clean = value.replace("_", " ")
    clean = "".join(ch if (ch.isalnum() or ch.isspace()) else " " for ch in clean)
    clean = " ".join(clean.split())
    return clean.strip().lower()


def resolve_device(device_arg: str = "auto") -> torch.device:
    if device_arg == "auto":
        return torch.device("cuda" if torch.cuda.is_available() else "cpu")
    return torch.device(device_arg)


def build_arg_parser() -> argparse.ArgumentParser:
    script_dir = Path(__file__).parent
    parser = argparse.ArgumentParser(description="Evaluate character classifier outputs.")
    parser.add_argument("--input_dir", required=True, help="Directory containing generated images.")
    parser.add_argument("--output_dir", help="Directory where metrics will be stored.")
    parser.add_argument("--unlearn", required=True, help="JSON list of characters expected to be unlearned.")
    parser.add_argument("--retain", required=True, help="JSON list of characters expected to be retained.")
    parser.add_argument(
        "--classifier_path",
        default=str(script_dir / DEFAULT_MODEL),
        help="Path to the classifier checkpoint.",
    )
    parser.add_argument(
        "--labels_path",
        default=str(script_dir / DEFAULT_LABELS),
        help="Path to labels.csv.",
    )
    parser.add_argument("--batch_size", type=int, default=64, help="Batch size for evaluation.")
    parser.add_argument("--num_workers", type=int, default=4, help="Number of dataloader workers.")
    parser.add_argument("--image_size", type=int, default=224, help="Image resolution expected by the classifier.")
    parser.add_argument("--device", default="auto", help="Torch device string (e.g. cuda:0).")
    parser.add_argument(
        "--extensions",
        nargs="+",
        default=list(SUPPORTED_EXTENSIONS),
        help="Image extensions to scan within the input directory.",
    )
    return parser


def parse_cli_args() -> argparse.Namespace:
    return build_arg_parser().parse_args()


def parse_args() -> argparse.Namespace:
    # Backward-compatible alias for older code importing parse_args().
    return parse_cli_args()


def load_label_metadata(labels_path: Path) -> tuple[list[str], Dict[str, str]]:
    entries: Dict[int, str] = {}
    with labels_path.open("r", encoding="utf-8") as handle:
        reader = csv.DictReader(handle)
        for row in reader:
            idx = int(row["id"])
            concept = row["concept"].strip()
            entries[idx] = concept

    max_idx = max(entries) if entries else -1
    idx_to_label = [None] * (max_idx + 1)
    for idx, label in entries.items():
        idx_to_label[idx] = label

    alias_lookup: Dict[str, str] = {}
    for label in idx_to_label:
        if not label:
            continue
        candidates = {label, label.split("(")[0]}
        for candidate in candidates:
            key = normalize_name(candidate)
            if key and key not in alias_lookup:
                alias_lookup[key] = label
    return idx_to_label, alias_lookup


def resolve_via_substring(entry: str, idx_to_label: Sequence[str]) -> str | None:
    target = normalize_name(entry)
    matches = []
    for label in idx_to_label:
        if not label:
            continue
        if target in normalize_name(label):
            matches.append(label)
    if len(matches) == 1:
        return matches[0]
    if len(matches) > 1:
        raise ValueError(f"Ambiguous character '{entry}'. Matches: {matches}")
    return None


def parse_character_arg(raw_value: str, alias_lookup: Dict[str, str], idx_to_label: Sequence[str]) -> List[str]:
    if raw_value is None:
        raise ValueError("Both --unlearn and --retain must be provided.")
    raw_value = raw_value.strip()
    if not raw_value:
        return []

    if raw_value.startswith("["):
        try:
            values = json.loads(raw_value)
        except json.JSONDecodeError as exc:
            raise ValueError(f"Unable to parse list: {raw_value}") from exc
    else:
        values = [raw_value]

    resolved: List[str] = []
    seen = set()
    for entry in values:
        if not isinstance(entry, str):
            entry = str(entry)
        canonical = alias_lookup.get(normalize_name(entry))
        if canonical is None:
            canonical = resolve_via_substring(entry, idx_to_label)
        if canonical is None:
            raise ValueError(f"Unknown character '{entry}'. Check labels.csv for valid names.")
        if canonical not in seen:
            resolved.append(canonical)
            seen.add(canonical)
    return resolved


def resolve_character_set_arg(
    value: Sequence[str] | str,
    alias_lookup: Dict[str, str],
    idx_to_label: Sequence[str],
) -> list[str]:
    if isinstance(value, str):
        return parse_character_arg(value, alias_lookup, idx_to_label)

    resolved: list[str] = []
    seen = set()
    for entry in value:
        canonical = alias_lookup.get(normalize_name(str(entry)))
        if canonical is None:
            canonical = resolve_via_substring(str(entry), idx_to_label)
        if canonical is None:
            raise ValueError(f"Unknown character '{entry}'. Check labels.csv for valid names.")
        if canonical not in seen:
            resolved.append(canonical)
            seen.add(canonical)
    return resolved


@dataclass
class ImageRecord:
    path: Path
    label_idx: int
    canonical_name: str
    prompt_idx: int
    sample_idx: int


def collect_image_records(
    input_dir: Path,
    alias_lookup: Dict[str, str],
    label_to_idx: Dict[str, int],
    extensions: Sequence[str],
) -> Dict[str, List[ImageRecord]]:
    records: Dict[str, List[ImageRecord]] = defaultdict(list)
    extensions = tuple(ext.lower() for ext in extensions)
    for root, _, files in os.walk(input_dir):
        for filename in files:
            suffix = os.path.splitext(filename)[1].lower()
            if suffix not in extensions:
                continue
            stem = os.path.splitext(filename)[0]
            try:
                char_token, prompt_token, sample_token = stem.rsplit("_", 2)
            except ValueError:
                continue

            canonical = alias_lookup.get(normalize_name(char_token))
            if canonical is None:
                continue
            label_idx = label_to_idx.get(canonical)
            if label_idx is None:
                continue

            try:
                prompt_idx = int(prompt_token)
                sample_idx = int(sample_token)
            except ValueError:
                continue

            records[canonical].append(
                ImageRecord(
                    path=Path(root) / filename,
                    label_idx=label_idx,
                    canonical_name=canonical,
                    prompt_idx=prompt_idx,
                    sample_idx=sample_idx,
                )
            )
    return records


class CharacterDataset(Dataset):
    def __init__(self, records: Sequence[ImageRecord], image_size: int):
        self.records = list(records)
        self.transform = transforms.Compose(
            [
                transforms.Resize((image_size, image_size)),
                transforms.ToTensor(),
            ]
        )

    def __len__(self) -> int:
        return len(self.records)

    def __getitem__(self, index: int):
        record = self.records[index]
        image = Image.open(record.path).convert("RGB")
        return self.transform(image), record.label_idx, record.canonical_name


def load_classifier(checkpoint_path: Path, device: torch.device) -> tuple[nn.Module, int]:
    print(f"Loading classifier weights from: {checkpoint_path}")
    state_dict = load_state_dict(checkpoint_path, device)
    fc_weight = state_dict.get("fc.weight")
    if fc_weight is None:
        raise RuntimeError("Checkpoint missing fc.weight; cannot infer num_classes.")
    num_classes = fc_weight.shape[0]
    model = models.resnet50(weights=None)
    model.fc = nn.Linear(model.fc.in_features, num_classes)
    model.load_state_dict(state_dict)
    model.to(device)
    return model, num_classes


def load_state_dict(checkpoint_path: Path, device: torch.device) -> dict:
    try:
        return torch.load(checkpoint_path, map_location=device)
    except RuntimeError as exc:
        if "failed finding central directory" not in str(exc):
            raise
        print("Checkpoint archive appears truncated; rebuilding ZIP structure...")
        repaired = rebuild_pytorch_zip_archive(checkpoint_path)
        try:
            return torch.load(repaired, map_location=device)
        except RuntimeError as repaired_exc:
            raise RuntimeError(
                "Checkpoint is missing required tensor shards. Please re-download "
                f"{checkpoint_path} (original error: {repaired_exc})."
            ) from repaired_exc


def rebuild_pytorch_zip_archive(checkpoint_path: Path) -> io.BytesIO:
    data = checkpoint_path.read_bytes()
    entries: List[tuple[str, bytes]] = []
    offset = 0
    signature = b"PK\x03\x04"
    while offset + len(signature) <= len(data) and data[offset : offset + 4] == signature:
        (
            _version,
            flag,
            compression,
            _mtime,
            _mdate,
            crc,
            comp_size,
            uncomp_size,
            name_len,
            extra_len,
        ) = struct.unpack("<HHHHHIIIHH", data[offset + 4 : offset + 30])
        name_start = offset + 30
        name_end = name_start + name_len
        name_bytes = data[name_start:name_end]
        try:
            name = name_bytes.decode("utf-8")
        except UnicodeDecodeError:
            name = name_bytes.decode("latin-1")
        offset = name_end + extra_len
        data_start = offset
        if flag & 0x8:
            data_end, crc, comp_size, uncomp_size, offset = locate_descriptor(data, data_start)
        else:
            data_end = data_start + comp_size
            offset = data_end
        payload = data[data_start:data_end]
        payload = decode_zip_payload(payload, compression)
        entries.append((name, payload))
        offset = max(offset, data_end)

    if not entries:
        raise RuntimeError("Failed to rebuild classifier checkpoint: no entries found.")
    print(f"Recovered {len(entries)} files from checkpoint archive.")

    buffer = io.BytesIO()
    root_prefix = entries[0][0].split("/", 1)[0] if entries and "/" in entries[0][0] else ""
    version_name = f"{root_prefix}/version" if root_prefix else "version"
    if version_name not in {name for name, _ in entries}:
        entries.append((version_name, b"3\n"))
    with zipfile.ZipFile(buffer, "w", compression=zipfile.ZIP_STORED) as zf:
        for name, payload in entries:
            zf.writestr(name, payload)
    buffer.seek(0)
    print("Successfully rebuilt ZIP archive in-memory.")
    return buffer


def locate_descriptor(data: bytes, data_start: int) -> tuple[int, int, int, int, int]:
    signature = b"PK\x07\x08"
    search_pos = data_start
    while True:
        idx = data.find(signature, search_pos)
        if idx == -1:
            next_header = find_next_local_header(data, data_start)
            data_end = next_header if next_header != -1 else len(data)
            comp_size = data_end - data_start
            return data_end, 0, comp_size, comp_size, data_end
        comp_size = int.from_bytes(data[idx + 8 : idx + 12], "little")
        data_end = idx
        if comp_size == data_end - data_start:
            crc = int.from_bytes(data[idx + 4 : idx + 8], "little")
            uncomp_size = int.from_bytes(data[idx + 12 : idx + 16], "little")
            return data_end, crc, comp_size, uncomp_size, idx + 16
        search_pos = idx + 1


def find_next_local_header(data: bytes, start: int) -> int:
    return data.find(b"PK\x03\x04", start + 4)


def decode_zip_payload(payload: bytes, compression: int) -> bytes:
    if compression == 0:
        return payload
    if compression == 8:
        decompressor = zlib.decompressobj(-zlib.MAX_WBITS)
        data = decompressor.decompress(payload)
        data += decompressor.flush()
        return data
    raise RuntimeError(f"Unsupported compression method {compression} in checkpoint archive.")


def evaluate(
    model: nn.Module,
    data_loader: DataLoader,
    idx_to_label: Sequence[str],
    device: torch.device,
) -> Dict[str, dict]:
    stats: Dict[str, dict] = {}
    model.eval()
    with torch.no_grad():
        for images, targets, names in tqdm(data_loader, desc="Evaluating Characters", unit="batch"):
            images = images.to(device)
            targets = targets.to(device)
            logits = model(images)
            predictions = logits.argmax(dim=1).cpu().tolist()
            target_list = targets.cpu().tolist()
            if isinstance(names, tuple):
                names = list(names)
            for name, target_idx, pred_idx in zip(names, target_list, predictions):
                if name not in stats:
                    stats[name] = {"total": 0, "correct": 0, "misclassified": Counter()}
                stats[name]["total"] += 1
                if target_idx == pred_idx:
                    stats[name]["correct"] += 1
                else:
                    pred_name = idx_to_label[pred_idx] if pred_idx < len(idx_to_label) else f"Class_{pred_idx}"
                    stats[name]["misclassified"][pred_name] += 1
    return stats


def write_results_excel(results: dict, output_dir: Path) -> Path:
    unlearn = results["unlearn"]
    retain = results["retain"]
    if len(unlearn) == 1:
        run_name = unlearn[0].replace(" ", "_")
    elif len(unlearn) > 1:
        run_name = f"Thru{unlearn[-1].replace(' ', '_')}"
    elif retain:
        run_name = retain[0].replace(" ", "_")
    else:
        run_name = "CharacterEval"

    xlsx_path = output_dir / f"{run_name}_metrics.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Metrics"

    header = ["Metric", run_name]
    ws.append(header)
    rows: List[List[str | float]] = []

    for character in unlearn:
        rows.append([f"UA: {character}", results["ua"][character]])
    for character in retain:
        rows.append([f"RA: {character}", results["ra"][character]])
    rows.append(["Avg UA", results["avg_ua"]])
    rows.append(["Avg RA", results["avg_ra"]])

    for row in rows:
        ws.append(row)

    default_font = Font(name="Roboto", size=12)
    bold_font = Font(name="Roboto", size=12, bold=True)
    default_align = Alignment(vertical="center")
    center_align = Alignment(horizontal="center", vertical="center")

    num_rows = len(rows) + 1
    num_cols = len(header)
    for row_idx in range(1, num_rows + 1):
        for col_idx in range(1, num_cols + 1):
            cell = ws[f"{get_column_letter(col_idx)}{row_idx}"]
            cell.font = default_font
            cell.alignment = default_align if not (col_idx > 1 and row_idx > 1) else center_align
    for row_idx in range(num_rows - 1, num_rows + 1):
        for col_idx in range(1, num_cols + 1):
            ws[f"{get_column_letter(col_idx)}{row_idx}"].font = bold_font
    for col_idx in range(2, num_cols + 1):
        col_letter = get_column_letter(col_idx)
        for row_idx in range(2, num_rows + 1):
            cell = ws[f"{col_letter}{row_idx}"]
            if isinstance(cell.value, (int, float)):
                cell.number_format = "0.00%"

    wb.save(xlsx_path)
    print(f"Excel file saved at: {xlsx_path}")
    return xlsx_path


def _select_character_records(
    records_by_char: Dict[str, List[ImageRecord]],
    character_subset: Sequence[str],
) -> tuple[list[ImageRecord], list[str]]:
    selected_records: List[ImageRecord] = []
    missing_from_dir: List[str] = []
    print(f"Total image(s) found: '{sum(len(v) for v in records_by_char.values())}'")
    for character in character_subset:
        char_records = records_by_char.get(character, [])
        print(f"Found {len(char_records)} image(s) for '{character}'")
        if not char_records:
            missing_from_dir.append(character)
        selected_records.extend(char_records)
    return selected_records, missing_from_dir


def _build_results_dict(
    stats: Dict[str, dict],
    unlearn: Sequence[str],
    retain: Sequence[str],
    input_dir: str,
    output_dir: str | None,
    missing_from_dir: Sequence[str],
) -> dict:
    character_subset = list(unlearn) + list(retain)
    total_images = sum(stats.get(character, {}).get("total", 0) for character in character_subset)
    results = {
        "unlearn": list(unlearn),
        "retain": list(retain),
        "input_dir": input_dir,
        "output_dir": output_dir,
        "total_images": total_images,
        "avg_ua": 0.0,
        "avg_ra": 0.0,
        "ua": {character: 0.0 for character in unlearn},
        "ra": {character: 0.0 for character in retain},
        "acc": {},
        "acc_adj": {},
        "misclassified": {},
        "missing_characters": list(missing_from_dir),
    }

    for character in character_subset:
        char_stats = stats.get(character, {"total": 0, "correct": 0, "misclassified": Counter()})
        total = char_stats["total"]
        acc = (char_stats["correct"] / total) if total else 0.0
        results["acc"][character] = acc
        results["acc_adj"][character] = acc
        results["misclassified"][character] = dict(char_stats["misclassified"])

    if unlearn:
        for character in unlearn:
            results["ua"][character] = 1.0 - results["acc_adj"][character]
            print(f"UA {character}: {results['ua'][character]:.2%}")
        results["avg_ua"] = sum(results["ua"].values()) / len(unlearn)

    if retain:
        for character in retain:
            results["ra"][character] = results["acc_adj"][character]
            print(f"RA {character}: {results['ra'][character]:.2%}")
        results["avg_ra"] = sum(results["ra"].values()) / len(retain)

    for tag, avg in (("UA", results["avg_ua"]), ("RA", results["avg_ra"])):
        print(f"Average {tag}: {avg:.2%}")
    return results


def _save_results_json(results: dict, output_dir: Path, filename: str) -> Path:
    output_dir.mkdir(parents=True, exist_ok=True)
    out_path = output_dir / filename
    with out_path.open("w", encoding="utf-8") as handle:
        json.dump(results, handle, indent=2)
    print(f"Saved metrics to {out_path}")
    return out_path


def evaluate_character_metrics(
    input_dir: str | Path,
    unlearn: Sequence[str] | str,
    retain: Sequence[str] | str,
    output_dir: str | Path | None = None,
    classifier_path: str | Path | None = None,
    labels_path: str | Path | None = None,
    batch_size: int = 64,
    num_workers: int = 4,
    image_size: int = 224,
    device: str = "auto",
    extensions: Sequence[str] = SUPPORTED_EXTENSIONS,
    results_json_name: str = "results.json",
    write_excel: bool = True,
) -> dict[str, Any]:
    script_dir = Path(__file__).parent
    input_dir = Path(input_dir)
    if not input_dir.exists():
        raise FileNotFoundError(f"Input directory not found: {input_dir}")

    labels_path = Path(labels_path) if labels_path is not None else script_dir / DEFAULT_LABELS
    if not labels_path.exists():
        raise FileNotFoundError(f"labels.csv not found at {labels_path}")

    classifier_path = Path(classifier_path) if classifier_path is not None else script_dir / DEFAULT_MODEL
    if not classifier_path.exists():
        raise FileNotFoundError(f"Classifier checkpoint not found: {classifier_path}")

    idx_to_label, alias_lookup = load_label_metadata(labels_path)
    label_to_idx = {label: idx for idx, label in enumerate(idx_to_label) if label}
    print(f"Found '{len(label_to_idx)}' labels")

    unlearn_list = resolve_character_set_arg(unlearn, alias_lookup, idx_to_label)
    retain_list = resolve_character_set_arg(retain, alias_lookup, idx_to_label)
    character_subset = unlearn_list + retain_list
    if not character_subset:
        raise ValueError("At least one character must be provided via unlearn/retain.")

    records_by_char = collect_image_records(
        input_dir=input_dir,
        alias_lookup=alias_lookup,
        label_to_idx=label_to_idx,
        extensions=[ext.lower() for ext in extensions],
    )
    selected_records, missing_from_dir = _select_character_records(records_by_char, character_subset)
    if missing_from_dir:
        print("Warning: no samples were found for the following characters:")
        for missing in missing_from_dir:
            print(f"  - {missing}")
    if not selected_records:
        raise RuntimeError("No matching images were found for the requested characters.")

    torch_device = resolve_device(device)
    dataset = CharacterDataset(selected_records, image_size=image_size)
    data_loader = DataLoader(
        dataset,
        batch_size=batch_size,
        shuffle=False,
        num_workers=num_workers,
        pin_memory=torch_device.type == "cuda",
    )

    model, checkpoint_classes = load_classifier(checkpoint_path=classifier_path, device=torch_device)
    if len(idx_to_label) < checkpoint_classes:
        idx_to_label.extend([f"Class_{idx}" for idx in range(len(idx_to_label), checkpoint_classes)])

    stats = evaluate(model, data_loader, idx_to_label, torch_device)
    output_dir_str = str(output_dir) if output_dir is not None else None
    results = _build_results_dict(
        stats=stats,
        unlearn=unlearn_list,
        retain=retain_list,
        input_dir=str(input_dir),
        output_dir=output_dir_str,
        missing_from_dir=missing_from_dir,
    )

    results_json_path = None
    xlsx_path = None
    if output_dir is not None:
        output_dir_path = Path(output_dir)
        results_json_path = _save_results_json(results, output_dir_path, results_json_name)
        if write_excel:
            xlsx_path = write_results_excel(results, output_dir_path)

    return {
        "results": results,
        "results_json_path": str(results_json_path) if results_json_path else None,
        "xlsx_path": str(xlsx_path) if xlsx_path else None,
    }


def run_evaluation_from_args(args: argparse.Namespace) -> dict[str, Any]:
    return evaluate_character_metrics(
        input_dir=args.input_dir,
        output_dir=args.output_dir,
        unlearn=args.unlearn,
        retain=args.retain,
        classifier_path=args.classifier_path,
        labels_path=args.labels_path,
        batch_size=args.batch_size,
        num_workers=args.num_workers,
        image_size=args.image_size,
        device=args.device,
        extensions=args.extensions,
    )


def main() -> None:
    args = parse_cli_args()
    run_evaluation_from_args(args)


if __name__ == "__main__":
    main()
