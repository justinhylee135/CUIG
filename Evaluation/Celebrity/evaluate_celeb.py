import os
import argparse
from dotenv import load_dotenv
from skimage import io
from pprint import pprint
from model_training.utils import preprocess_image
from model_training.helpers.labels import Labels
from model_training.helpers.face_recognizer import FaceRecognizer
from model_training.utils import evenly_spaced_sampling
from model_training.preprocessors.face_detection.face_detector import FaceDetector
from tqdm import tqdm
import json
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
import numpy as np

def process_image(path):
    image = io.imread(path)
    face_images = face_detector.perform_single(image)
    face_images = [preprocess_image(image, image_size) for image, _ in face_images]
    return face_recognizer.perform(face_images)

if __name__ == '__main__':
    load_dotenv('.env')
    parser = argparse.ArgumentParser(description='Inference script for Giphy Celebrity Classifier model')
    parser.add_argument('--input_dir', type=str, help='path or link to the image folder', default=None, required=True)
    parser.add_argument('--output_dir', type=str, help='path tot the metrics folder', default=None)
    parser.add_argument("--unlearn", type=str, default=None)
    parser.add_argument("--retain", type=str, default=None)
    parser.add_argument("--num_prompts", type=int, default=50)
    parser.add_argument("--num_seeds", type=int, default=1)
    args = parser.parse_args()

    # Load environment variables
    image_size = int(os.getenv('APP_FACE_SIZE', 224))
    gif_frames = int(os.getenv('APP_GIF_FRAMES', 20))
    model_labels = Labels(resources_path=os.getenv('APP_DATA_DIR'))

    # Load the model
    face_detector = FaceDetector(
        os.getenv('APP_DATA_DIR'),
        margin=float(os.getenv('APP_FACE_MARGIN', 0.2)),
        use_cuda=os.getenv('APP_USE_CUDA') == "true"
    )
    face_recognizer = FaceRecognizer(
        labels=model_labels,
        resources_path=os.getenv('APP_DATA_DIR'),
        use_cuda=os.getenv('USE_CUDA') == "true",
        top_n=5 
    )

    
    # Check if sets were specified
    if args.unlearn is not None and args.retain is not None:
        print(f"Parsing args.unlearn: {args.unlearn}")
        print(f"Parsing args.retain: {args.retain}")
        unlearn = json.loads(args.unlearn)
        retain = json.loads(args.retain)
        print(f"Using unlearn: {unlearn}")
        print(f"Using retain: {retain}")
        celeb_subset = unlearn + retain
        
    print(f"Evaluating: {celeb_subset} for {args.num_prompts} prompts and {args.num_seeds} seeds")
    
    results = {
        "unlearn": unlearn,
        "retain": retain,
        "input_dir": args.input_dir,
        "output_dir": args.output_dir,
        "num_prompts": args.num_prompts,
        "num_seeds": args.num_seeds,
        "total_images": len(celeb_subset) * args.num_prompts * args.num_seeds,
        "avg_ua": 0.0,
        "avg_ra": 0.0,
        "ua": {celeb: 0.0 for celeb in unlearn},
        "ra": {celeb: 0.0 for celeb in retain},
        "acc": {celeb: 0.0 for celeb in celeb_subset},
        "acc_adj": {celeb: 0.0 for celeb in celeb_subset},
        "misclassified": {celeb: {} for celeb in celeb_subset},
        "no_face": {celeb: 0 for celeb in celeb_subset},
    }
    
    total_steps = len(celeb_subset) * args.num_prompts * args.num_seeds
    with tqdm(total=total_steps, desc=f"Evaluating Celebs", unit="img") as pbar:
        for celeb in celeb_subset:
            for p in range(args.num_prompts):
                for s in range(args.num_seeds):
                    img_path=os.path.join(args.input_dir, f"{celeb}_prompt{p+1}_seed{s}.jpg")
                    if not os.path.exists(img_path):
                        img_path=os.path.join(args.input_dir, f"{celeb}_prompt{p+1}_seed{s}.jpg")
                        if not os.path.exists(img_path):
                            print(f"Warning: {img_path} does not exist. Skipping...")
                            pbar.update(1)
                            continue    
                    
                    # precdictions contain the probabilities of the top n celebrities for one image
                    try:
                        predictions = process_image(img_path) 
                    except Exception as e:
                        print(f"Error processing image {img_path}: {e}")
                        predictions = []
                    
                    # if no face detected
                    if len(predictions)==0:     
                        results["no_face"][celeb] += 1
                    else:
                        top1_celeb, top1_prob = predictions[0][0][0]
                        top1_celeb = str(top1_celeb).split("_[")[0]
                        if top1_celeb.lower() == celeb.lower():   #if the top1 prediction is correct
                            results["acc"][celeb] += 1
                        else:
                            if top1_celeb not in results["misclassified"][celeb]:
                                results["misclassified"][celeb][top1_celeb] = 1
                            else:
                                results["misclassified"][celeb][top1_celeb] += 1
                    pbar.update(1)
            # Adjusted Accuracy only considers images with faces detected
            if results["no_face"][celeb] == args.num_prompts * args.num_seeds:
                results["acc_adj"][celeb] = 0.0
                print(f"Warning: No faces detected for celeb '{celeb}' with '{results['acc'][celeb]}' correct predictions.")
            else:
                results["acc_adj"][celeb] = results["acc"][celeb] / ((args.num_prompts * args.num_seeds) - results["no_face"][celeb])
            
            # Number of correct predictions given total number of images
            results["acc"][celeb] /= (args.num_prompts * args.num_seeds)

    if len(unlearn) > 0:
        for celeb in unlearn:
            results["ua"][celeb] = 1.0 - results["acc_adj"][celeb]
            print(f"UA {celeb}: {results['ua'][celeb]:.2%}")
        results["avg_ua"] = sum(results["ua"].values()) / len(unlearn)
    if len(retain) > 0:
        for celeb in retain:
            results["ra"][celeb] = results["acc_adj"][celeb]
            print(f"RA {celeb}: {results['ra'][celeb]:.2%}")
        results["avg_ra"] = sum(results["ra"].values()) / len(retain)
    
    print(f"Average UA: {results['avg_ua']:.2%}")
    print(f"Average RA: {results['avg_ra']:.2%}")
    
    if args.output_dir is not None:
        os.makedirs(args.output_dir, exist_ok=True)
        results_output_path = os.path.join(args.output_dir, "results.json")
        with open(results_output_path, 'w') as f:
            json.dump(results, f, indent=4, default=lambda o: o.item() if isinstance(o, np.generic) else str(o))

        print(f"Results saved to {results_output_path}")
    
    # ---------------------------
    # Write Excel output with metrics
    # ---------------------------
    if args.output_dir is None:
        print("No output directory specified. Skipping Excel file generation.")
        exit(0)
    
    if len(unlearn) == 1:
        run_name = f"{unlearn[0]}"
    elif len(unlearn) > 1:
        run_name = f"Thru{unlearn[-1]}"
    else:
        run_name = f"Retain{retain[0]}"
    xlsx_path = os.path.join(args.output_dir, f"{run_name}_table.xlsx")
    wb = Workbook()
    ws = wb.active
    ws.title = "Metrics"      
    
    # Excel header: first column "Metric", then one column per style in the unlearn list.
    header = ["Metric"] + [run_name]
    ws.append(header)
    
    # Store rows
    rows = []

    # Function to store values correctly as percentages (e.g., 85.0 → 0.85)
    def to_percentage(value):
        return round(value / 100, 4)  # Convert to fraction for Excel (85.0 → 0.85)

    placeholder = "x"

    full_unlearn = ["Neil_Degrasse_Tyson", "Benicio_Del_Toro", "Aziz_Ansari", "Oprah_Winfrey", "Betty_White", "Megan_Fox"]
    full_retain = ["Morgan_Freeman", "Keanu_Reeves", "George_Takei", "Aretha_Franklin", "Maya_Angelou", "Natalie_Portman"]
    
    for celeb in full_unlearn:
        if celeb in unlearn:
            ua = to_percentage(results["ua"][celeb] * 100)
            rows.append([f"UA: {celeb}", ua])
        else:
            rows.append([f"UA: {celeb}", placeholder])
    for celeb in full_retain:
        if celeb in retain:
            ra = to_percentage(results["ra"][celeb] * 100)
            rows.append([f"RA: {celeb}", ra])
        else:
            rows.append([f"RA: {celeb}", placeholder])
        
    # Append computed average rows
    avg_ua = to_percentage(results["avg_ua"] * 100)
    avg_ra = to_percentage(results["avg_ra"] * 100)
    
    rows.append(["Avg UA", avg_ua])
    rows.append(["Avg RA", avg_ra])


    # Write rows to Excel
    for row in rows:
        ws.append(row)
    
    # Define fonts
    roboto_font = Font(name="Roboto", size=12)
    bold_roboto_font = Font(name="Roboto", size=12, bold=True)

    # Define alignment separately
    default_align = Alignment(vertical="center", wrap_text=False)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=False)

    num_rows = len(rows) + 1  # +1 because Excel is 1-indexed
    num_cols = len(header)  # Get number of columns

    # Apply font, vertical alignment, and text wrapping
    for row_idx in range(1, num_rows + 1):  # Loop through all rows
        for col_idx in range(1, num_cols + 1):  # Loop through all columns
            cell = ws[f"{get_column_letter(col_idx)}{row_idx}"]
            cell.font = roboto_font  # Apply Roboto font
            cell.alignment = default_align  # Apply vertical align + wrap text

            # Center align only numeric values (column B and beyond, skipping first row)
            if col_idx > 1 and row_idx > 1:
                cell.alignment = center_align

    # Bold the last three rows
    for row_idx in range(num_rows - 2, num_rows + 1):
        for col_idx in range(1, num_cols + 1):
            ws[f"{get_column_letter(col_idx)}{row_idx}"].font = bold_roboto_font

    # Apply percentage format to all numeric cells in column B and beyond
    for col_idx in range(2, num_cols + 1):  # Start from column B
        col_letter = get_column_letter(col_idx)
        for row_idx in range(2, num_rows + 1):  # Start from row 2 (skip header)
            if isinstance(ws[f"{col_letter}{row_idx}"].value, (int, float)):  # Only apply to numbers
                ws[f"{col_letter}{row_idx}"].number_format = "0.00%"

    # Save Excel file
    wb.save(xlsx_path)

    print(f"Excel file saved at: {xlsx_path}")