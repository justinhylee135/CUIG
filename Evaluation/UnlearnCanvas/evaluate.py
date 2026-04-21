# Standard Library
import argparse
import ast
import json
import os
from pathlib import Path
from typing import Any, Sequence

# Third Party
import numpy as np
import timm
import torch
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from PIL import Image
from torchvision import transforms
from tqdm import tqdm

torch.hub.set_dir("cache")

# Local
try:
    # Package-style import for reuse from other Python modules.
    from .constants import (
        OBJECTS_AVAILABLE,
        STYLES_AVAILABLE,
        XLSX_ALL_RETAIN_OBJECTS,
        XLSX_ALL_RETAIN_STYLES,
        XLSX_ALL_UNLEARN_OBJECTS,
        XLSX_ALL_UNLEARN_STYLES,
    )
except ImportError:
    # Script-style fallback for direct `python evaluate.py ...` execution.
    from constants import (
        OBJECTS_AVAILABLE,
        STYLES_AVAILABLE,
        XLSX_ALL_RETAIN_OBJECTS,
        XLSX_ALL_RETAIN_STYLES,
        XLSX_ALL_UNLEARN_OBJECTS,
        XLSX_ALL_UNLEARN_STYLES,
    )


def build_arg_parser() -> argparse.ArgumentParser:
    # Split parser creation from execution so tools/tests can reuse it.
    parser = argparse.ArgumentParser()

    # Directory arguments
    parser.add_argument("--input_dir", type=str, help="Directory storing images to evaluate", required=True)
    parser.add_argument("--output_dir", type=str, help="Directory to store results output.", required=True)
    parser.add_argument(
        "--seed",
        type=int,
        nargs="+",
        help="Seeds that were used for sampling",
        default=[188, 288, 588, 688, 888],
    )

    # Evaluation Settings
    parser.add_argument(
        "--eval_classifier_dir",
        type=str,
        help="Directory that holds style and object classifier",
        required=True,
    )
    parser.add_argument("--device", type=str, default="cuda:0")

    # Specify concepts that were unlearned and retained
    parser.add_argument("--unlearn", type=str, help="Set of concepts that were unlearned", default=None)
    parser.add_argument("--retain", type=str, help="Set of in-domain concepts to retain", default=None)
    parser.add_argument("--cross_retain", type=str, help="Set of cross-domain concepts to retain", default=None)
    return parser


def parse_cli_args() -> argparse.Namespace:
    # Thin wrapper to preserve a simple CLI entry point.
    return build_arg_parser().parse_args()


def resolve_device(device_arg: str | torch.device) -> torch.device:
    # Accept both CLI strings and torch.device values from in-process callers.
    if not torch.cuda.is_available():
        return torch.device("cpu")

    if isinstance(device_arg, torch.device):
        if device_arg.type != "cuda":
            return torch.device(device_arg.type)
        device_index = device_arg.index
        if device_index is None:
            device_index = torch.cuda.current_device()
    else:
        device_text = str(device_arg)
        if device_text == "cpu":
            return torch.device("cpu")
        if not device_text.startswith("cuda"):
            return torch.device(device_text)
        if ":" in device_text:
            _, index_text = device_text.split(":", 1)
            device_index = int(index_text)
        else:
            device_index = torch.cuda.current_device()

    torch.cuda.set_device(device_index)
    return torch.device(f"cuda:{device_index}")


def _parse_list_arg(raw_value: str, name: str) -> list[str]:
    # CLI passes concept sets as Python literals, e.g. "['Bears', 'Cats']".
    parsed = ast.literal_eval(raw_value)
    if not isinstance(parsed, (list, tuple)):
        raise ValueError(f"Argument '{name}' must be a list-like literal. Got: {raw_value}")
    return [str(x) for x in parsed]


def resolve_evaluation_subsets(
    unlearn_arg: str | None,
    retain_arg: str | None,
    cross_retain_arg: str | None,
) -> dict[str, Any]:
    # Start with the full benchmark by default; callers can override with subsets.
    styles_available_subset = list(STYLES_AVAILABLE)
    objects_available_subset = list(OBJECTS_AVAILABLE)

    # Require all three subset arguments together to avoid ambiguous summary metrics.
    provided = [unlearn_arg is not None, retain_arg is not None, cross_retain_arg is not None]
    if any(provided) and not all(provided):
        raise ValueError("If specifying evaluation subsets, provide all of --unlearn, --retain, and --cross_retain.")

    unlearn = None
    retain = None
    cross_retain = None
    is_style_unlearn = None

    if all(provided):
        # Parse user-specified concept groups used for UA / IRA / CRA reporting.
        unlearn = _parse_list_arg(unlearn_arg, "unlearn")
        retain = _parse_list_arg(retain_arg, "retain")
        cross_retain = _parse_list_arg(cross_retain_arg, "cross_retain")
        print(f"Using Unlearn: {unlearn}")
        print(f"Using Retain: {retain}")
        print(f"Using Cross Retain: {cross_retain}")

        if not unlearn:
            raise ValueError("'unlearn' list is empty.")

        # Determine if we're unlearning style or object to determine what in-domain and cross-domain retain set
        is_style_unlearn = unlearn[0] in STYLES_AVAILABLE
        if is_style_unlearn:
            styles_available_subset = unlearn + retain
            objects_available_subset = cross_retain
        else:
            styles_available_subset = cross_retain
            objects_available_subset = unlearn + retain

    return {
        "styles_subset": styles_available_subset,
        "objects_subset": objects_available_subset,
        "unlearn": unlearn,
        "retain": retain,
        "cross_retain": cross_retain,
        "is_style_unlearn": is_style_unlearn,
    }


def build_image_transform():
    # Match the classifier input preprocessing expected by the ViT classifiers.
    return transforms.Compose(
        [
            transforms.Resize((224, 224)),
            transforms.ToTensor(),
            transforms.Normalize([0.5], [0.5]),
        ]
    )


def load_task_classifier(task: str, eval_classifier_dir: str, device: torch.device):
    # Recreate the classifier backbone/head exactly as done during classifier training.
    model = timm.create_model("vit_large_patch16_224.augreg_in21k", pretrained=True).to(device)
    num_classes = len(STYLES_AVAILABLE) if task == "style" else len(OBJECTS_AVAILABLE)
    model.head = torch.nn.Linear(1024, num_classes).to(device)

    # Each task has its own classifier checkpoint (`style_classifier.pth` or `object_classifier.pth`).
    classifier_path = Path(eval_classifier_dir) / f"{task}_classifier.pth"
    model.load_state_dict(torch.load(classifier_path, map_location=device, weights_only=False)["model_state_dict"])
    model.eval()
    return model


def init_task_results(
    task: str,
    input_dir: str,
    styles_available_subset: Sequence[str],
    objects_available_subset: Sequence[str],
) -> dict[str, Any]:
    # The output schema 
    if task == "style":
        return {
            "input_dir": input_dir,
            "acc": {style: 0.0 for style in styles_available_subset},
            "misclassified": {
                style: {other_style: 0 for other_style in STYLES_AVAILABLE}
                for style in styles_available_subset
            },
            "image_specific": {},
            "eval_count": {style: 0 for style in styles_available_subset},
            "pred_loss": {style: 0.0 for style in styles_available_subset},
            "loss": {style: 0.0 for style in styles_available_subset},
        }

    return {
        "input_dir": input_dir,
        "acc": {obj: 0.0 for obj in objects_available_subset},
        "misclassified": {
            obj: {other_object: 0 for other_object in OBJECTS_AVAILABLE}
            for obj in objects_available_subset
        },
        "image_specific": {},
        "eval_count": {obj: 0 for obj in objects_available_subset},
        "pred_loss": {obj: 0.0 for obj in objects_available_subset},
        "loss": {obj: 0.0 for obj in objects_available_subset},
    }


def _record_prediction(
    results: dict[str, Any],
    key: str,
    loss: torch.Tensor,
    pred_loss: torch.Tensor,
    pred_success: torch.Tensor,
    predicted_name: str,
    img_path: str,
) -> None:
    # Centralized bookkeeping so style/object loops stay compact and consistent.
    results["loss"][key] += loss.item()
    results["pred_loss"][key] += pred_loss.item()
    results["acc"][key] += pred_success.item()
    results["eval_count"][key] += 1
    results["misclassified"][key][predicted_name] += 1
    short_img_path = os.path.splitext(img_path.split("images/")[1])[0]
    results["image_specific"][short_img_path] = predicted_name


def evaluate_style_task(
    model,
    input_dir: str,
    device: torch.device,
    image_transform,
    styles_available_subset: Sequence[str],
    objects_available_subset: Sequence[str],
    seed_list: Sequence[int],
    results: dict[str, Any],
    progress,
) -> None:
    # Style classifier predicts style labels while object varies as cross-domain context.
    for test_style in styles_available_subset:
        style_label = STYLES_AVAILABLE.index(test_style)
        for seed in seed_list:
            for test_object in objects_available_subset:
                img_path = os.path.join(input_dir, f"{test_style}_{test_object}_seed{seed}.jpg")
                # Missing images can happen if sampling was interrupted; skip but keep progress aligned.
                if not os.path.exists(img_path):
                    print(f"Warning: Image {img_path} not found. Skipping...")
                    progress.update(1)
                    continue

                # Evaluate one image at a time to preserve the original script's semantics and logging.
                image = Image.open(img_path)
                target_image = image_transform(image).unsqueeze(0).to(device)
                with torch.no_grad():
                    res = model(target_image)
                    pred_label = torch.argmax(res)
                    label = torch.tensor([style_label]).to(device)
                    loss = torch.nn.functional.cross_entropy(res, label)
                    res_softmax = torch.nn.functional.softmax(res, dim=1)
                    pred_loss = res_softmax[0][style_label]
                    pred_success = (pred_label == style_label).sum()

                predicted_name = STYLES_AVAILABLE[pred_label.item()]
                _record_prediction(results, test_style, loss, pred_loss, pred_success, predicted_name, img_path)
                progress.update(1)


def evaluate_object_task(
    model,
    input_dir: str,
    device: torch.device,
    image_transform,
    styles_available_subset: Sequence[str],
    objects_available_subset: Sequence[str],
    seed_list: Sequence[int],
    results: dict[str, Any],
    progress,
) -> None:
    # Object classifier predicts object labels while style varies as cross-domain context.
    for test_style in styles_available_subset:
        for seed in seed_list:
            for test_object in objects_available_subset:
                object_label = OBJECTS_AVAILABLE.index(test_object)
                img_path = os.path.join(input_dir, f"{test_style}_{test_object}_seed{seed}.jpg")
                # Missing images can happen if sampling was interrupted; skip but keep progress aligned.
                if not os.path.exists(img_path):
                    print(f"Warning: Image {img_path} not found. Skipping...")
                    progress.update(1)
                    continue

                # Evaluate one image at a time to preserve the original script's semantics and logging.
                image = Image.open(img_path)
                target_image = image_transform(image).unsqueeze(0).to(device)
                with torch.no_grad():
                    res = model(target_image)
                    pred_label = torch.argmax(res)
                    label = torch.tensor([object_label]).to(device)
                    loss = torch.nn.functional.cross_entropy(res, label)
                    res_softmax = torch.nn.functional.softmax(res, dim=1)
                    pred_loss = res_softmax[0][object_label]
                    pred_success = (pred_label == object_label).sum()

                predicted_name = OBJECTS_AVAILABLE[pred_label.item()]
                _record_prediction(results, test_object, loss, pred_loss, pred_success, predicted_name, img_path)
                progress.update(1)


def normalize_and_prune_results(results: dict[str, Any]) -> None:
    # Convert running counts into percentages and clean zero-count confusion matrix entries.
    for key in results["acc"]:
        count = results["eval_count"][key]
        results["acc"][key] = results["acc"][key] / count if count > 0 else 0.0
        results["acc"][key] = round(100 * results["acc"][key], 2)

    for concept in results["misclassified"]:
        for other_concept in list(results["misclassified"][concept].keys()):
            if results["misclassified"][concept][other_concept] == 0:
                del results["misclassified"][concept][other_concept]


def save_results_json(results: dict[str, Any], output_path: str) -> None:
    # JSON serialization helper shared by style/object task evaluation.
    with open(output_path, "w") as handle:
        # Some values may still be tensors; convert them before writing to JSON.
        serializable_results = {
            k: {sk: (float(sv) if isinstance(sv, torch.Tensor) else sv) for sk, sv in v.items()}
            if isinstance(v, dict)
            else v
            for k, v in results.items()
        }
        json.dump(serializable_results, handle, indent=4)


def evaluate_single_task(
    task: str,
    input_dir: str,
    output_dir: str,
    seed_list: Sequence[int],
    styles_available_subset: Sequence[str],
    objects_available_subset: Sequence[str],
    eval_classifier_dir: str,
    device: torch.device,
    image_transform,
) -> dict[str, Any]:
    # Evaluate one task ("style" or "object") and emit its JSON report.
    output_path = os.path.join(output_dir, f"{task}_results.json")
    model = load_task_classifier(task, eval_classifier_dir, device)
    results = init_task_results(task, input_dir, styles_available_subset, objects_available_subset)

    # Progress count is the full Cartesian product of subsets and seeds.
    total_steps = len(styles_available_subset) * len(seed_list) * len(objects_available_subset)
    with tqdm(total=total_steps, desc=f"Evaluating {task}", unit="img") as pbar:
        if task == "style":
            evaluate_style_task(
                model=model,
                input_dir=input_dir,
                device=device,
                image_transform=image_transform,
                styles_available_subset=styles_available_subset,
                objects_available_subset=objects_available_subset,
                seed_list=seed_list,
                results=results,
                progress=pbar,
            )
        elif task == "object":
            evaluate_object_task(
                model=model,
                input_dir=input_dir,
                device=device,
                image_transform=image_transform,
                styles_available_subset=styles_available_subset,
                objects_available_subset=objects_available_subset,
                seed_list=seed_list,
                results=results,
                progress=pbar,
            )
        else:
            raise ValueError(f"Unknown task: {task}")

    normalize_and_prune_results(results)
    save_results_json(results, output_path)
    return results


def to_percentage(value: float) -> float:
    # Excel stores percentages as fractional values (e.g., 0.85 for 85%).
    return round(value / 100, 4)


def compute_unlearncanvas_summary(
    results_style: dict[str, Any],
    results_object: dict[str, Any],
    unlearn: Sequence[str],
    retain: Sequence[str],
    cross_retain: Sequence[str],
    is_style_unlearn: bool,
) -> tuple[dict[str, Any], dict[str, Any], dict[str, Any], float, float, float]:
    # "Primary" task is the same domain as the unlearned concept (style or object).
    primary_results = results_style if is_style_unlearn else results_object
    secondary_results = results_object if is_style_unlearn else results_style

    # UA is reported per-unlearn concept as (100 - accuracy).
    for item in unlearn:
        accuracy = 100.0 - primary_results["acc"][item]
        print(f"Unlearn accuracy for {item}: {accuracy:.2f}%")

    # Aggregate metrics used in the paper/README.
    unlearn_accuracy_avg = 100.0 - np.mean([primary_results["acc"][item] for item in unlearn])
    retain_accuracy = np.mean([primary_results["acc"][item] for item in retain])
    cross_retain_accuracy = np.mean([secondary_results["acc"][item] for item in cross_retain])

    # Summary schema is preserved for backward compatibility with prior outputs.
    summary = {
        "UA": round(unlearn_accuracy_avg, 2),
        "IRA": round(retain_accuracy, 2),
        "CRA": round(cross_retain_accuracy, 2),
        "unlearn": {},
        "retain": {},
        "cross_retain": {},
    }
    for concept in unlearn:
        summary["unlearn"][concept] = 100.0 - primary_results["acc"][concept]
    for concept in retain:
        summary["retain"][concept] = primary_results["acc"][concept]
    for concept in cross_retain:
        summary["cross_retain"][concept] = secondary_results["acc"][concept]

    return summary, primary_results, secondary_results, float(unlearn_accuracy_avg), float(retain_accuracy), float(cross_retain_accuracy)


def save_summary_json(output_dir: str, summary: dict[str, Any]) -> None:
    # Write compact summary metrics to a stable filename for downstream tooling.
    with open(os.path.join(output_dir, "summary.json"), "w") as handle:
        json.dump(summary, handle, indent=4)


def print_summary_metrics(unlearn_accuracy_avg: float, retain_accuracy: float, cross_retain_accuracy: float) -> None:
    # Console output mirrors the previous script for easy human inspection.
    print(f"Unlearn accuracy (average): {unlearn_accuracy_avg:.2f}%")
    print(f"Retain accuracy (average): {retain_accuracy:.2f}%")
    print(f"Cross retain accuracy (average): {cross_retain_accuracy:.2f}%")


def write_excel_report(
    output_dir: str,
    unlearn: Sequence[str],
    retain: Sequence[str],
    cross_retain: Sequence[str],
    primary_results: dict[str, Any],
    secondary_results: dict[str, Any],
    unlearn_accuracy_avg: float,
    retain_accuracy: float,
    cross_retain_accuracy: float,
) -> str:
    # Build the same paper-style Excel sheet the original script generated.
    if unlearn[0] in XLSX_ALL_UNLEARN_STYLES:
        full_unlearn = XLSX_ALL_UNLEARN_STYLES
        full_retain = XLSX_ALL_RETAIN_STYLES
        full_cross_retain = XLSX_ALL_RETAIN_OBJECTS
    else:
        full_unlearn = XLSX_ALL_UNLEARN_OBJECTS
        full_retain = XLSX_ALL_RETAIN_OBJECTS
        full_cross_retain = XLSX_ALL_RETAIN_STYLES

    run_name = f"{unlearn[0]}" if len(unlearn) == 1 else f"Thru{unlearn[-1]}"
    xlsx_path = os.path.join(output_dir, f"{run_name}_table.xlsx")

    # Create workbook and a single metrics worksheet.
    wb = Workbook()
    ws = wb.active
    ws.title = "Metrics"

    header = ["Metric", run_name]
    ws.append(header)

    rows = []
    placeholder = "x"

    # Populate per-concept UA/IRA/CRA rows, using placeholders for concepts not in this run.
    for style in full_unlearn:
        if style in unlearn:
            ua = to_percentage(100.0 - primary_results["acc"][style])
            rows.append([f"UA: {style}", ua])
        else:
            rows.append([f"UA: {style}", placeholder])

    for concept in full_retain:
        if concept in retain:
            ira = to_percentage(primary_results["acc"][concept])
            rows.append([f"IRA: {concept}", ira])
        else:
            rows.append([f"IRA: {concept}", placeholder])

    for concept in full_cross_retain:
        if concept in cross_retain:
            cra = to_percentage(secondary_results["acc"][concept])
            rows.append([f"CRA: {concept}", cra])
        else:
            rows.append([f"CRA: {concept}", placeholder])

    rows.append(["Avg UA", to_percentage(unlearn_accuracy_avg)])
    rows.append(["Avg IRA", to_percentage(retain_accuracy)])
    rows.append(["Avg CRA", to_percentage(cross_retain_accuracy)])

    # Write computed rows to the sheet in one pass.
    for row in rows:
        ws.append(row)

    # Apply lightweight styling for readability.
    roboto_font = Font(name="Roboto", size=12)
    bold_roboto_font = Font(name="Roboto", size=12, bold=True)
    default_align = Alignment(vertical="center", wrap_text=False)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=False)

    num_rows = len(rows) + 1
    num_cols = len(header)

    # Apply default formatting and center-align numeric columns.
    for row_idx in range(1, num_rows + 1):
        for col_idx in range(1, num_cols + 1):
            cell = ws[f"{get_column_letter(col_idx)}{row_idx}"]
            cell.font = roboto_font
            cell.alignment = default_align
            if col_idx > 1 and row_idx > 1:
                cell.alignment = center_align

    for row_idx in range(num_rows - 2, num_rows + 1):
        for col_idx in range(1, num_cols + 1):
            ws[f"{get_column_letter(col_idx)}{row_idx}"].font = bold_roboto_font

    # Convert Excel cells to percentage formatting for numeric metric cells only.
    for col_idx in range(2, num_cols + 1):
        col_letter = get_column_letter(col_idx)
        for row_idx in range(2, num_rows + 1):
            if isinstance(ws[f"{col_letter}{row_idx}"].value, (int, float)):
                ws[f"{col_letter}{row_idx}"].number_format = "0.00%"

    wb.save(xlsx_path)
    print(f"Excel file saved at: {xlsx_path}")
    return xlsx_path


def evaluate_unlearncanvas(
    input_dir: str,
    output_dir: str,
    eval_classifier_dir: str,
    device: str = "cuda:0",
    seed_list: Sequence[int] = (188, 288, 588, 688, 888),
    unlearn: Sequence[str] | None = None,
    retain: Sequence[str] | None = None,
    cross_retain: Sequence[str] | None = None,
    styles_available_subset: Sequence[str] | None = None,
    objects_available_subset: Sequence[str] | None = None,
    is_style_unlearn: bool | None = None,
) -> dict[str, Any]:
    # Main reusable evaluation API: evaluates both tasks and optionally computes summary exports.
    os.makedirs(output_dir, exist_ok=True)
    torch_device = resolve_device(device)
    image_transform = build_image_transform()

    # If subsets are not provided, evaluate the full benchmark.
    styles_available_subset = list(STYLES_AVAILABLE if styles_available_subset is None else styles_available_subset)
    objects_available_subset = list(OBJECTS_AVAILABLE if objects_available_subset is None else objects_available_subset)

    # Evaluate style and object tasks separately, matching the original script outputs.
    results_style = evaluate_single_task(
        task="style",
        input_dir=input_dir,
        output_dir=output_dir,
        seed_list=seed_list,
        styles_available_subset=styles_available_subset,
        objects_available_subset=objects_available_subset,
        eval_classifier_dir=eval_classifier_dir,
        device=torch_device,
        image_transform=image_transform,
    )
    results_object = evaluate_single_task(
        task="object",
        input_dir=input_dir,
        output_dir=output_dir,
        seed_list=seed_list,
        styles_available_subset=styles_available_subset,
        objects_available_subset=objects_available_subset,
        eval_classifier_dir=eval_classifier_dir,
        device=torch_device,
        image_transform=image_transform,
    )

    output = {
        "results_style": results_style,
        "results_object": results_object,
        "summary": None,
        "xlsx_path": None,
    }

    # Summary/Excel requires explicit unlearn/retain/cross-retain semantics.
    if unlearn is None or retain is None or cross_retain is None or is_style_unlearn is None:
        print("Subset summary/export skipped because unlearn/retain/cross_retain were not provided.")
        return output

    summary, primary_results, secondary_results, ua_avg, ira_avg, cra_avg = compute_unlearncanvas_summary(
        results_style=results_style,
        results_object=results_object,
        unlearn=unlearn,
        retain=retain,
        cross_retain=cross_retain,
        is_style_unlearn=is_style_unlearn,
    )
    save_summary_json(output_dir, summary)
    print_summary_metrics(ua_avg, ira_avg, cra_avg)
    xlsx_path = write_excel_report(
        output_dir=output_dir,
        unlearn=unlearn,
        retain=retain,
        cross_retain=cross_retain,
        primary_results=primary_results,
        secondary_results=secondary_results,
        unlearn_accuracy_avg=ua_avg,
        retain_accuracy=ira_avg,
        cross_retain_accuracy=cra_avg,
    )

    output["summary"] = summary
    output["xlsx_path"] = xlsx_path
    return output


def run_evaluation_from_args(args: argparse.Namespace) -> dict[str, Any]:
    # CLI adapter: parse subset arguments and then call the reusable evaluator.
    subset_cfg = resolve_evaluation_subsets(args.unlearn, args.retain, args.cross_retain)
    return evaluate_unlearncanvas(
        input_dir=args.input_dir,
        output_dir=args.output_dir,
        eval_classifier_dir=args.eval_classifier_dir,
        device=args.device,
        seed_list=args.seed,
        unlearn=subset_cfg["unlearn"],
        retain=subset_cfg["retain"],
        cross_retain=subset_cfg["cross_retain"],
        styles_available_subset=subset_cfg["styles_subset"],
        objects_available_subset=subset_cfg["objects_subset"],
        is_style_unlearn=subset_cfg["is_style_unlearn"],
    )


def main() -> None:
    # CLI entry point kept minimal for easier reuse and testing.
    args = parse_cli_args()
    run_evaluation_from_args(args)


if __name__ == "__main__":
    main()
